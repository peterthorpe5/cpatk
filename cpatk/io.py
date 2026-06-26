"""Input and output helpers for CPATK.

The package intentionally writes tab-separated, Parquet, Excel and HTML files.
Comma-separated output is not used by CPATK writing helpers.
"""

from __future__ import annotations

import html
import logging
from pathlib import Path
from typing import Dict, Mapping, Optional, Sequence, Union

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

TableLike = Union[pd.DataFrame, Mapping[str, pd.DataFrame]]

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
DEFAULT_EXCEL_PREVIEW_ROWS = 100_000


def is_ignored_sidecar_path(*, path: Union[str, Path]) -> bool:
    """Return whether a path should be ignored during table discovery.

    Parameters
    ----------
    path:
        Candidate path to evaluate.

    Returns
    -------
    bool
        True when the file or any parent component is a hidden/system sidecar.

    Notes
    -----
    macOS AppleDouble files such as ``._table.tsv`` can be created when data
    are copied from a Mac. They are not real tables and often contain binary
    metadata bytes, so CPATK must ignore them during folder discovery.
    """
    candidate = Path(path)
    for part in candidate.parts:
        if part in {".", ".."}:
            continue
        if part.startswith("._") or part.startswith(".") or part.startswith("~$"):
            return True
    return False


def read_table(
    *,
    path: Union[str, Path],
    sheet_name: Optional[Union[str, int]] = None,
    logger: Optional[logging.Logger] = None,
) -> pd.DataFrame:
    """Read a table from TSV, CSV, Excel or Parquet.

    Parameters
    ----------
    path:
        Input table path.
    sheet_name:
        Sheet name or index for Excel input.
    logger:
        Optional logger.

    Returns
    -------
    pandas.DataFrame
        Loaded table.
    """
    path = Path(path)
    suffixes = "".join(path.suffixes).lower()
    if logger is not None:
        logger.info("Reading table: %s", path)

    if suffixes.endswith(".parquet"):
        return pd.read_parquet(path=path)
    if suffixes.endswith(".tsv") or suffixes.endswith(".tsv.gz"):
        return pd.read_csv(filepath_or_buffer=path, sep="\t", encoding="utf-8-sig", encoding_errors="replace")
    if suffixes.endswith(".csv") or suffixes.endswith(".csv.gz"):
        return pd.read_csv(filepath_or_buffer=path, encoding="utf-8-sig", encoding_errors="replace", low_memory=False)
    if suffixes.endswith(".xlsx") or suffixes.endswith(".xls"):
        return pd.read_excel(io=path, sheet_name=sheet_name or 0)

    raise ValueError(f"Unsupported input table format: {path}")


def write_table(
    *,
    data_frame: pd.DataFrame,
    path: Union[str, Path],
    index: bool = False,
    logger: Optional[logging.Logger] = None,
) -> Path:
    """Write a table to TSV, TSV.GZ, Parquet or Excel.

    Parameters
    ----------
    data_frame:
        Data frame to write.
    path:
        Output path. Supported suffixes are ``.tsv``, ``.tsv.gz``,
        ``.parquet`` and ``.xlsx``.
    index:
        Whether to write the data-frame index.
    logger:
        Optional logger.

    Returns
    -------
    pathlib.Path
        Written output path.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffixes = "".join(path.suffixes).lower()
    if logger is not None:
        logger.info("Writing table: %s", path)

    if suffixes.endswith(".parquet"):
        data_frame.to_parquet(path=path, index=index)
    elif suffixes.endswith(".tsv") or suffixes.endswith(".tsv.gz"):
        data_frame.to_csv(path_or_buf=path, sep="\t", index=index)
    elif suffixes.endswith(".xlsx"):
        write_excel_workbook(
            tables={"Sheet1": data_frame},
            path=path,
            logger=logger,
        )
    else:
        raise ValueError(f"Unsupported output table format: {path}")
    return path


def sanitise_sheet_name(*, sheet_name: str) -> str:
    """Return a valid Excel sheet name no longer than 31 characters.

    Parameters
    ----------
    sheet_name:
        Proposed sheet name.

    Returns
    -------
    str
        Sanitised sheet name.
    """
    invalid = set("[]:*?/\\")
    cleaned = "".join("_" if character in invalid else character for character in sheet_name)
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def _unique_sheet_name(*, sheet_name: str, used_names: set[str]) -> str:
    """Return a unique, Excel-safe sheet name."""
    safe_name = sanitise_sheet_name(sheet_name=sheet_name)
    base_name = safe_name
    counter = 1
    while safe_name in used_names:
        suffix = f"_{counter}"
        safe_name = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(safe_name)
    return safe_name


def _excel_safe_preview(
    *,
    table: pd.DataFrame,
    max_preview_rows: int,
) -> tuple[pd.DataFrame, dict[str, Union[str, int, bool]]]:
    """Return an Excel-safe preview and an audit record for a table.

    Excel is used by CPATK as a readable summary format. Lossless exports are
    the TSV/Parquet files written beside each workbook. Very large tables are
    therefore previewed in Excel rather than causing the whole workflow to fail.
    """
    original_rows = int(table.shape[0])
    original_columns = int(table.shape[1])
    max_data_rows = max(0, min(max_preview_rows, EXCEL_MAX_ROWS - 1))
    max_columns = EXCEL_MAX_COLUMNS
    truncated_rows = original_rows > max_data_rows
    truncated_columns = original_columns > max_columns
    preview = table.iloc[:max_data_rows, :max_columns].copy()
    record: dict[str, Union[str, int, bool]] = {
        "original_rows": original_rows,
        "original_columns": original_columns,
        "rows_written_to_excel": int(preview.shape[0]),
        "columns_written_to_excel": int(preview.shape[1]),
        "truncated_rows": bool(truncated_rows),
        "truncated_columns": bool(truncated_columns),
        "reason": "previewed_for_excel_limit" if truncated_rows or truncated_columns else "complete",
    }
    return preview, record


def write_excel_workbook(
    *,
    tables: Mapping[str, pd.DataFrame],
    path: Union[str, Path],
    logger: Optional[logging.Logger] = None,
    max_preview_rows: int = DEFAULT_EXCEL_PREVIEW_ROWS,
) -> Path:
    """Write a formatted Excel workbook with one sheet per table.

    Parameters
    ----------
    tables:
        Mapping of sheet names to data frames.
    path:
        Output workbook path.
    logger:
        Optional logger.
    max_preview_rows:
        Maximum number of data rows to write for an individual Excel sheet.
        TSV/Parquet outputs remain the lossless data exports; Excel sheets are
        readable summaries and are previewed when tables are too large.

    Returns
    -------
    pathlib.Path
        Written workbook path.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if logger is not None:
        logger.info("Writing formatted Excel workbook: %s", path)

    export_notes: list[dict[str, Union[str, int, bool]]] = []
    with pd.ExcelWriter(path=path, engine="openpyxl") as writer:
        used_names: set[str] = set()
        for sheet_name, table in tables.items():
            safe_name = _unique_sheet_name(sheet_name=sheet_name, used_names=used_names)
            preview, note = _excel_safe_preview(table=table, max_preview_rows=max_preview_rows)
            note["requested_sheet_name"] = str(sheet_name)
            note["excel_sheet_name"] = safe_name
            export_notes.append(note)
            if note["reason"] != "complete" and logger is not None:
                logger.warning(
                    "Excel sheet '%s' is too large for a readable workbook; "
                    "writing preview rows=%s columns=%s from original rows=%s columns=%s.",
                    sheet_name,
                    note["rows_written_to_excel"],
                    note["columns_written_to_excel"],
                    note["original_rows"],
                    note["original_columns"],
                )
            preview.to_excel(excel_writer=writer, sheet_name=safe_name, index=False)
            worksheet = writer.sheets[safe_name]
            format_worksheet(worksheet=worksheet, data_frame=preview)

        if export_notes:
            notes_name = _unique_sheet_name(sheet_name="Excel_export_notes", used_names=used_names)
            notes_table = pd.DataFrame.from_records(export_notes)
            notes_table.to_excel(excel_writer=writer, sheet_name=notes_name, index=False)
            worksheet = writer.sheets[notes_name]
            format_worksheet(worksheet=worksheet, data_frame=notes_table)
    return path


def format_worksheet(*, worksheet, data_frame: pd.DataFrame) -> None:
    """Apply readable formatting to an openpyxl worksheet.

    Parameters
    ----------
    worksheet:
        Openpyxl worksheet object.
    data_frame:
        Data frame written to the worksheet.
    """
    header_fill = PatternFill(fill_type="solid", fgColor="ECEFF1")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(data_frame.columns, start=1):
        values = data_frame[column_name].astype(str).head(200).tolist()
        max_length = max([len(str(column_name)), *[len(value) for value in values]])
        width = min(max(max_length + 2, 10), 45)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)


def data_frame_to_html_table(
    *,
    data_frame: pd.DataFrame,
    max_rows: int = 20,
) -> str:
    """Convert a data frame to a small HTML table.

    Parameters
    ----------
    data_frame:
        Data frame to render.
    max_rows:
        Maximum number of rows to show.

    Returns
    -------
    str
        HTML table string.
    """
    shown = data_frame.head(n=max_rows).copy()
    header = "".join(f"<th>{html.escape(str(column))}</th>" for column in shown.columns)
    rows = []
    for _, row in shown.iterrows():
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in row.tolist())
        rows.append(f"<tr>{cells}</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody></table>"


def list_supported_tables(*, input_dir: Union[str, Path]) -> pd.DataFrame:
    """List recognised table files in a directory.

    Parameters
    ----------
    input_dir:
        Directory to scan.

    Returns
    -------
    pandas.DataFrame
        Inventory of candidate table files.
    """
    input_dir = Path(input_dir)
    patterns = ["*.tsv", "*.tsv.gz", "*.csv", "*.csv.gz", "*.parquet", "*.xlsx"]
    records = []
    for pattern in patterns:
        for path in sorted(input_dir.glob(pattern)):
            if is_ignored_sidecar_path(path=path):
                continue
            records.append(
                {
                    "path": str(path),
                    "file_name": path.name,
                    "size_bytes": path.stat().st_size,
                    "suffixes": "".join(path.suffixes),
                }
            )
    return pd.DataFrame.from_records(records)

"""Regression tests for CPATK v0.2.16 reporting/workbook hardening."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cpatk.clipn_adapter import ClipnAdapterConfig, clean_impute_and_scale_aligned
from cpatk.io import read_table, write_excel_workbook, write_table
from cpatk.reporting import discover_plot_paths, make_html_report


class TestExcelWorkbookHardeningV0216(unittest.TestCase):
    """Tests for Excel previewing of oversized summary tables."""

    def test_large_excel_sheet_is_previewed_and_audited(self) -> None:
        """Large tables should not crash workbook writing."""
        table = pd.DataFrame({"a": range(25), "b": range(25)})
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.xlsx"
            write_excel_workbook(
                tables={"large_table": table},
                path=path,
                max_preview_rows=10,
            )
            workbook = load_workbook(filename=path, read_only=True)
            self.assertIn("large_table", workbook.sheetnames)
            self.assertIn("Excel_export_notes", workbook.sheetnames)
            self.assertEqual(workbook["large_table"].max_row, 11)
            notes = workbook["Excel_export_notes"]
            headers = [cell.value for cell in next(notes.iter_rows(min_row=1, max_row=1))]
            values = [cell.value for cell in next(notes.iter_rows(min_row=2, max_row=2))]
            record = dict(zip(headers, values))
            self.assertEqual(record["original_rows"], 25)
            self.assertEqual(record["rows_written_to_excel"], 10)
            self.assertTrue(record["truncated_rows"])


class TestHtmlReportV0216(unittest.TestCase):
    """Tests for plainer, more useful HTML reports."""

    def test_report_uses_summary_heading_and_links_tables(self) -> None:
        """Reports should use the plainer Summary heading."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            table_path = write_table(
                data_frame=pd.DataFrame({"item": ["rows"], "value": [3]}),
                path=tmp / "summary.tsv",
            )
            output = tmp / "report.html"
            table = read_table(path=table_path)
            make_html_report(
                title="Test report",
                output_path=output,
                summary_tables={"Summary table": table},
                table_paths={"Summary table": table_path},
                narrative="Plain report.",
            )
            html = output.read_text(encoding="utf-8")
            self.assertIn("<h2>Summary</h2>", html)
            self.assertNotIn("Executive", html)
            self.assertIn("Result map", html)
            self.assertIn("Open full table", html)

    def test_auto_plot_discovery_finds_svg_outputs(self) -> None:
        """Plot discovery should find common figure outputs."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            svg = tmp / "plots" / "embedding.svg"
            svg.parent.mkdir()
            svg.write_text("<svg></svg>", encoding="utf-8")
            discovered = discover_plot_paths(root_dir=tmp, output_html=tmp / "report.html")
            self.assertIn(svg, discovered)


class TestClipnFilteringMessageV0216(unittest.TestCase):
    """Tests for clearer CLIPn strict-zero filtering failures."""

    def test_empty_after_any_zero_filter_has_clear_message(self) -> None:
        """Strict any-zero filtering should fail before sklearn scaler errors."""
        aligned = {
            "a": pd.DataFrame({"f1": [1.0, 0.0], "f2": [0.0, 2.0]}),
            "b": pd.DataFrame({"f1": [3.0, 0.0], "f2": [0.0, 4.0]}),
        }
        metadata = {
            "a": pd.DataFrame({"id": ["a1", "a2"]}),
            "b": pd.DataFrame({"id": ["b1", "b2"]}),
        }
        config = ClipnAdapterConfig(drop_rows_with_any_zero=True, scaling_method="none")
        with self.assertRaisesRegex(ValueError, "removed every sample"):
            clean_impute_and_scale_aligned(
                aligned=aligned,
                metadata=metadata,
                config=config,
            )


if __name__ == "__main__":
    unittest.main()
